"""
Order-Manager für LegoLAS.

Importiert und exportiert Auftragslisten als Excel-Dateien (.xlsx).

Spaltenformat der Excel-Vorlage:
  A: Teilenummer (part_num)
  B: Name (name)
  C: Anzahl (required)
  D: Behälter (container, 1–6)

Prioritätsregel: Behälter 1 hat höchste Priorität,
                 Behälter 6 ist die Aussortierschublade.
"""

import logging
import os
from typing import List, Optional, Tuple

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
    logger.warning("openpyxl nicht installiert – Excel-Import/Export nicht verfügbar.")


# Standardfarben für Excel-Export
_HEADER_FILL = "2E4057"
_ROW_FILLS = ["FFFFFF", "F0F4F8"]


class OrderManager:
    """
    Excel-basierter Auftragsimporter/-exporter.

    Parameter
    ---------
    orders_dir : str
        Verzeichnis für Auftragsdateien.
    exports_dir : str
        Verzeichnis für exportierte Dateien.
    """

    def __init__(self, orders_dir: str, exports_dir: str):
        self.orders_dir = orders_dir
        self.exports_dir = exports_dir
        os.makedirs(orders_dir, exist_ok=True)
        os.makedirs(exports_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    def import_order(self, filepath: str) -> Tuple[str, List[Tuple[str, int, int]]]:
        """
        Liest eine Auftragsliste aus einer Excel-Datei.

        Rückgabe
        --------
        (order_name, items)
        items = [(part_num, container, required), ...]
        """
        if not _OPENPYXL:
            raise RuntimeError("openpyxl ist nicht installiert.")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        order_name = os.path.splitext(os.path.basename(filepath))[0]
        items: List[Tuple[str, int, int]] = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1 and self._is_header(row):
                continue  # Kopfzeile überspringen
            if not row or row[0] is None:
                continue

            part_num = str(row[0]).strip()
            required = int(row[2]) if len(row) > 2 and row[2] else 1
            container = int(row[3]) if len(row) > 3 and row[3] else 6
            container = max(1, min(6, container))
            items.append((part_num, container, required))

        wb.close()
        logger.info("Auftrag importiert: %s (%d Positionen)", order_name, len(items))
        return order_name, items

    @staticmethod
    def _is_header(row) -> bool:
        if not row or row[0] is None:
            return False
        return str(row[0]).lower() in {"teilenummer", "part_num", "part", "id"}

    # ------------------------------------------------------------------
    # Export: Auftragsliste
    # ------------------------------------------------------------------

    def export_order(self, order_name: str,
                     items: List[dict], filepath: str = None) -> str:
        """
        Exportiert eine Auftragsliste als Excel.

        ``items`` Listenelemente sind Dicts mit keys:
          part_num, name, container, required, fulfilled
        """
        if not _OPENPYXL:
            raise RuntimeError("openpyxl ist nicht installiert.")

        if filepath is None:
            safe_name = "".join(
                c for c in order_name if c.isalnum() or c in " _-"
            ).strip()
            filepath = os.path.join(
                self.exports_dir, f"{safe_name}_auftrag.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Auftrag"

        headers = ["Teilenummer", "Name", "Behälter",
                   "Soll", "Ist", "Offen"]
        self._write_header(ws, headers)

        for idx, item in enumerate(sorted(items, key=lambda x: x.get("container", 6))):
            fill_hex = _ROW_FILLS[idx % 2]
            fill = PatternFill("solid", fgColor=fill_hex)
            offen = max(0, item.get("required", 0) - item.get("fulfilled", 0))
            row_data = [
                item.get("part_num", ""),
                item.get("name", ""),
                item.get("container", ""),
                item.get("required", 0),
                item.get("fulfilled", 0),
                offen,
            ]
            row_idx = idx + 3
            ws.append(row_data)
            for cell in ws[row_idx]:
                cell.fill = fill

        self._auto_column_width(ws)
        wb.save(filepath)
        logger.info("Auftrag exportiert: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # Export: Fehlende Teile je Behälter
    # ------------------------------------------------------------------

    def export_missing_parts(self, order_name: str,
                              items: List[dict],
                              filepath: str = None) -> str:
        """
        Exportiert eine Liste der fehlenden Teile je Behälter.
        """
        if not _OPENPYXL:
            raise RuntimeError("openpyxl ist nicht installiert.")

        if filepath is None:
            safe_name = "".join(
                c for c in order_name if c.isalnum() or c in " _-"
            ).strip()
            filepath = os.path.join(
                self.exports_dir, f"{safe_name}_fehlend.xlsx")

        missing = [i for i in items
                   if i.get("fulfilled", 0) < i.get("required", 0)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Fehlende Teile"

        headers = ["Behälter", "Teilenummer", "Name", "Soll", "Ist", "Fehlt"]
        self._write_header(ws, headers)

        for idx, item in enumerate(sorted(missing,
                                          key=lambda x: x.get("container", 6))):
            fill_hex = _ROW_FILLS[idx % 2]
            fill = PatternFill("solid", fgColor=fill_hex)
            fehlt = item.get("required", 0) - item.get("fulfilled", 0)
            ws.append([
                item.get("container", ""),
                item.get("part_num", ""),
                item.get("name", ""),
                item.get("required", 0),
                item.get("fulfilled", 0),
                fehlt,
            ])
            for cell in ws[idx + 3]:
                cell.fill = fill

        self._auto_column_width(ws)
        wb.save(filepath)
        logger.info("Fehlende Teile exportiert: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # Export: Inventar
    # ------------------------------------------------------------------

    def export_inventory(self, inventory: List[dict],
                         filepath: str = None) -> str:
        """Exportiert den aktuellen Inventarbestand."""
        if not _OPENPYXL:
            raise RuntimeError("openpyxl ist nicht installiert.")
        if filepath is None:
            filepath = os.path.join(self.exports_dir, "inventar.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventar"

        headers = ["Teilenummer", "Name", "Behälter", "Anzahl", "Zuletzt aktualisiert"]
        self._write_header(ws, headers)

        for idx, item in enumerate(inventory):
            fill_hex = _ROW_FILLS[idx % 2]
            fill = PatternFill("solid", fgColor=fill_hex)
            ws.append([
                item.get("part_num", ""),
                item.get("name", ""),
                item.get("container", ""),
                item.get("count", 0),
                item.get("updated_at", ""),
            ])
            for cell in ws[idx + 3]:
                cell.fill = fill

        self._auto_column_width(ws)
        wb.save(filepath)
        logger.info("Inventar exportiert: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # Hilfsmethoden
    # ------------------------------------------------------------------

    def _write_header(self, ws, headers: List[str]):
        ws.append([""])  # Leerzeile oben für Ästhetik
        ws.append(headers)
        fill = PatternFill("solid", fgColor=_HEADER_FILL)
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[2]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_column_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
